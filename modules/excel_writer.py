"""
Excel Writer — Natijalarni Excel ga yozish
===========================================
financial_engine natijalarini data.xlsx shablonga yozadi.
Loans sheetini to'liq qayta yaratadi (Python hisoblangan qiymatlar).
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import shutil
import os
import logging

logger = logging.getLogger(__name__)

# Stillar
HEADER_FONT = Font(name="Arial", bold=True, size=9, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
DATA_FONT = Font(name="Arial", size=8)
BOLD_FONT = Font(name="Arial", bold=True, size=8)
TOTAL_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
IMTIYOZ_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
NUM_FMT = '#,##0'


def _style_cell(cell, font=None, fill=None, align=None, border=None, fmt=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if border: cell.border = border
    if fmt: cell.number_format = fmt


def write_excel_output(template_path: str, output_path: str, engine) -> str:
    """
    Excel shablonni nusxalab, engine natijalarini yozadi.
    1. Asosiy ma'lumotlarni birinchi sheetga.
    2. Loans sheetini to'liq qayta yaratadi.
    3. Boshqa sheetlarga qiymatlar yozadi.
    """
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)

    # === 1. Asosiy ma'lumotlar (birinchi sheet) ===
    ws = wb.worksheets[0]
    field_map = {
        "B16": "loyiha_nomi", "B18": "tashabbuskor", "B20": "manzil",
        "B22": "bank", "B24": "stir", "B26": "jshshir",
        "B28": "faoliyat", "B30": "mulk", "B32": "fio",
        "B34": "pasport", "B36": "berilgan_vaqti",
        "B38": "loyiha_qiymati", "B42": "oz_mablag",
        "B44": "kredit", "B46": "muddat", "B48": "imtiyoz", "B50": "foiz",
        "B56": "mahsulot", "B58": "hajm", "B60": "olchov", "B62": "narx",
        "B65": "direktor", "B67": "xodim", "B69": "yangi_xodim",
        "B72": "ishchi_oylik", "B74": "rahbar_oylik", "B76": "yangi_ishchi_oylik",
        "B81": "elektr", "B83": "gaz", "B85": "suv", "B87": "oqava",
    }
    ctx = engine.get_context()
    for cell_ref, field_name in field_map.items():
        val = ctx.get(field_name, "")
        ws[cell_ref] = val

    # Kredit summary
    kn = engine.kredit_natija
    ws["B90"] = kn.oylik_tolov
    ws["B91"] = kn.jami_tolov
    ws["B92"] = kn.jami_foiz

    # === 2. Loans sheet — to'liq qayta yaratish ===
    _write_loans_sheet(wb, engine)

    # === 3. Boshqa sheetlarga yillik qiymatlarni yozish ===
    _write_yearly_values(wb, engine)

    wb.save(output_path)
    logger.info(f"Excel saqlandi: {output_path}")
    return output_path


def _write_loans_sheet(wb, engine):
    """Loans sheetini tozalab, Python kredit jadvalini yozadi."""
    sheet_name = "Loans"
    if sheet_name in wb.sheetnames:
        # Eski ma'lumotlarni tozalash
        ws = wb[sheet_name]
        # Sarlavha va parametrlarni saqlab, jadval qismini tozalash
        for row in range(12, ws.max_row + 1):
            for col in range(1, 8):
                ws.cell(row=row, column=col).value = None
    else:
        ws = wb.create_sheet(sheet_name)

    # Parametrlar
    ws["C3"] = "Kredit muddati"
    ws["D3"] = engine.muddat_oy
    ws["E3"] = "oy"
    ws["C5"] = "Imtiyozli davr"
    ws["D5"] = engine.imtiyoz_oy
    ws["E5"] = "oy"
    ws["C6"] = "Foiz"
    ws["D6"] = engine.foiz / 100.0
    ws["E6"] = "foiz"
    ws["C9"] = "Kredit summasi"
    ws["D9"] = engine.kredit_summa
    ws["E9"] = "so'm"
    ws["F1"] = "8-9-ILOVA"

    # Jadval sarlavha
    headers = ["OYLAR", "ASOSIY QARZ TO'LOVI", "ASOSIY QARZ QOLDIG'I",
               "FOIZ TO'LOVLARI", "JAMI TO'LOV", "YIL"]
    for i, h in enumerate(headers, 2):
        cell = ws.cell(row=11, column=i, value=h)
        _style_cell(cell, HEADER_FONT, HEADER_FILL,
                    Alignment(horizontal="center", vertical="center", wrap_text=True),
                    THIN_BORDER)
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Ma'lumotlar (Python hisoblagan qiymatlar — VALUES, formulalar emas!)
    kn = engine.kredit_natija
    for idx, t in enumerate(kn.jadval):
        row = 12 + idx
        fill = IMTIYOZ_FILL if t.imtiyozli else None
        yil_num = (t.oy - 1) // 12 + 1

        vals = [t.oy, t.asosiy_qarz, t.qoldiq, t.foiz_tolov, t.oylik_tolov, f"{yil_num}-yil"]
        for i, v in enumerate(vals, 2):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill
            if isinstance(v, float):
                cell.number_format = NUM_FMT

    # JAMI qatori
    last_row = 12 + len(kn.jadval)
    ws.cell(row=last_row, column=2, value="JAMI").font = BOLD_FONT
    jami_vals = [None, kn.kredit_summa, 0, kn.jami_foiz, kn.jami_tolov, ""]
    for i, v in enumerate(jami_vals, 2):
        if v is not None:
            cell = ws.cell(row=last_row, column=i + 1, value=v)
            _style_cell(cell, BOLD_FONT, TOTAL_FILL,
                       Alignment(horizontal="center"), THIN_BORDER, NUM_FMT)


def _write_yearly_values(wb, engine):
    """Boshqa sheetlarga yillik qiymatlarni yozish (formulalar o'rniga VALUES)."""
    years = 7
    model_years = engine.MODEL_YEARS

    # Depreciate sheet
    if "Depreciate" in wb.sheetnames:
        ws = wb["Depreciate"]
        dep_list = engine.t_depreciation["data"]["yearly_list"]
        for y in range(years):
            col = 5 + y  # E=5, F=6, ..., K=11
            val = dep_list[y] if y < len(dep_list) else ""
            ws.cell(row=5, column=col, value=val)
            ws.cell(row=7, column=col, value=val)
            ws.cell(row=8, column=col, value=(val * (y + 1)) if val != "" else "")

    # ProdPlan revenue
    if "ProdPlan" in wb.sheetnames:
        ws = wb["ProdPlan"]
        revs = engine.t_prod_plan["data"]["yearly_revenue"]
        caps = engine.CAPACITIES
        for y in range(years):
            col = 3 + y  # C=3, D=4, ..., I=9
            ws.cell(row=24, column=col, value=caps[y] if y < len(caps) else "")
            ws.cell(row=58, column=col, value=revs[y] if y < len(revs) else "")

    # ProfLoss
    if "ProfLoss" in wb.sheetnames:
        ws = wb["ProfLoss"]
        for y in range(years):
            col = 3 + y
            if y < model_years:
                p = engine.yearly_pnl[y]
                ws.cell(row=4, column=col, value=p["daromad"])
                ws.cell(row=6, column=col, value=p["sof_daromad"])
                ws.cell(row=7, column=col, value=p["tannarx"])
                ws.cell(row=8, column=col, value=p["yalpi_foyda"])
                ws.cell(row=12, column=col, value=p["operatsion_foyda"])
                ws.cell(row=13, column=col, value=p["foiz"])
                ws.cell(row=14, column=col, value=p["ebt"])
                ws.cell(row=15, column=col, value=p["soliq"])
                ws.cell(row=17, column=col, value=p["sof_foyda"])
            else:
                for row_idx in [4, 6, 7, 8, 12, 13, 14, 15, 17]:
                    ws.cell(row=row_idx, column=col, value="")

    # CashFlow
    if "CashFlow" in wb.sheetnames:
        ws = wb["CashFlow"]
        ws.cell(row=15, column=3, value=-engine.loyiha_qiymati)
        for y in range(years):
            col = 4 + y  # D=4, ..., J=10
            if y < model_years:
                c = engine.yearly_cf[y]
                ws.cell(row=4, column=col, value=c["daromad"])
                ws.cell(row=7, column=col, value=c["sof_tushum"])
                ws.cell(row=8, column=col, value=c["tannarx"])
                ws.cell(row=9, column=col, value=c["yalpi_tushum"])
                ws.cell(row=14, column=col, value=c["operatsion_cf"])
                ws.cell(row=18, column=col, value=c["foiz"])
                ws.cell(row=19, column=col, value=c["soliq"])
            else:
                for row_idx in [4, 7, 8, 9, 14, 18, 19]:
                    ws.cell(row=row_idx, column=col, value="")

    # NPV
    if "npv" in wb.sheetnames:
        ws = wb["npv"]
        ind = engine.indicators
        ws.cell(row=6, column=3, value=-engine.loyiha_qiymati)
        r = engine.discount_rate / 100
        for y in range(years):
            if y < model_years:
                cf = engine.yearly_cf[y]["sof_cf"]
                ws.cell(row=7 + y, column=3, value=cf)
                disc = 1 / ((1 + r) ** (y + 1))
                ws.cell(row=7 + y, column=4, value=disc)
                ws.cell(row=7 + y, column=5, value=cf * disc)
            else:
                ws.cell(row=7 + y, column=3, value="")
                ws.cell(row=7 + y, column=4, value="")
                ws.cell(row=7 + y, column=5, value="")

        ws.cell(row=16, column=5, value=ind["npv"])
        ws.cell(row=18, column=5, value=ind["irr"])

    # Taxes
    if "Taxes" in wb.sheetnames:
        ws = wb["Taxes"]
        for y in range(years):
            col = 5 + y  # E=5, ..., K=11
            if y < model_years:
                tax_data = engine.yearly_taxes[y]
                if engine.soliq_turi == "mchj":
                    ws.cell(row=5, column=col, value=tax_data["qqs"])
                else:
                    ws.cell(row=4, column=col, value=tax_data["aylanma"])
                ws.cell(row=8, column=col, value=tax_data["jami"])
            else:
                ws.cell(row=4, column=col, value="")
                ws.cell(row=5, column=col, value="")
                ws.cell(row=8, column=col, value="")

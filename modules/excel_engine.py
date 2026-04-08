"""
Dinamik Excel Engine
====================
Kredit jadvalini muddat bo'yicha dinamik yaratish.
Moliyaviy ko'rsatkichlarni Excel'ga yozish.
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
import shutil
import os
from typing import Dict, Any, Optional
from .credit_calculator import KreditNatija
from .financial_analysis import full_financial_analysis


# Stillar
HEADER_FONT = Font(name='Times New Roman', bold=True, size=11, color='000000')
HEADER_FILL = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')
DATA_FONT = Font(name='Times New Roman', size=11)
BOLD_FONT = Font(name='Times New Roman', bold=True, size=11)
TOTAL_FILL = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
IMTIYOZ_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
NUM_FORMAT = '#,##0'
NUM_FORMAT_2 = '#,##0.00'


def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, num_fmt=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border
    if num_fmt: cell.number_format = num_fmt


def write_kredit_jadval(ws, start_row: int, kredit: KreditNatija) -> int:
    """Kredit to'lov jadvalini Excel sheet'ga yozish. Qaytaradi: keyingi bo'sh qator."""

    # Sarlavha
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    title_cell = ws.cell(row=start_row, column=1,
                         value=f"KREDIT TO'LOV JADVALI ({kredit.turi.upper()})")
    apply_cell_style(title_cell, Font(name='Times New Roman', bold=True, size=14, color='000000'),
                     alignment=Alignment(horizontal='center'))

    row = start_row + 1
    # Umumiy ma'lumot
    info = [
        ("Kredit summasi:", kredit.kredit_summa),
        ("Foiz stavkasi:", f"{kredit.foiz_stavka}%"),
        ("Muddat:", f"{kredit.muddat_oy} oy"),
        ("Imtiyozli davr:", f"{kredit.imtiyoz_oy} oy"),
    ]
    for label, val in info:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.font = DATA_FONT
        if isinstance(val, (int, float)):
            c.number_format = NUM_FORMAT
        row += 1

    row += 1

    # Jadval sarlavhalari
    headers = ["Oy", "Oylik to'lov", "Asosiy qarz", "Foiz to'lov", "Qoldiq", "Holat"]
    col_widths = [6, 18, 18, 18, 18, 12]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_cell_style(cell, HEADER_FONT, HEADER_FILL,
                         Alignment(horizontal='center', vertical='center'), THIN_BORDER)
        ws.column_dimensions[get_column_letter(i)].width = w

    row += 1

    # Ma'lumotlar
    for t in kredit.jadval:
        fill = IMTIYOZ_FILL if t.imtiyozli else None
        for i, val in enumerate([t.oy, t.oylik_tolov, t.asosiy_qarz,
                                  t.foiz_tolov, t.qoldiq,
                                  "Imtiyozli" if t.imtiyozli else "Asosiy"], 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if fill:
                cell.fill = fill
            if isinstance(val, float):
                cell.number_format = NUM_FORMAT
        row += 1

    # Jami
    row_jami = row
    ws.cell(row=row, column=1, value="JAMI").font = BOLD_FONT
    for i, val in enumerate([kredit.jami_tolov, kredit.kredit_summa,
                              kredit.jami_foiz, 0, ""], 2):
        cell = ws.cell(row=row, column=i, value=val)
        apply_cell_style(cell, BOLD_FONT, TOTAL_FILL, Alignment(horizontal='center'), THIN_BORDER)
        if isinstance(val, (int, float)):
            cell.number_format = NUM_FORMAT

    return row + 2


def write_financial_analysis(ws, start_row: int, analysis: dict) -> int:
    """Moliyaviy tahlil natijalarini yozish."""

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=5)
    title = ws.cell(row=start_row, column=1, value="MOLIYAVIY TAHLIL NATIJALARI")
    apply_cell_style(title, Font(name='Times New Roman', bold=True, size=14, color='000000'),
                     alignment=Alignment(horizontal='center'))
    row = start_row + 2

    # Ko'rsatkichlar
    indicators = [
        ("NPV (Sof joriy qiymat)", analysis["npv"], "so'm"),
        ("IRR (Ichki daromadlilik)", analysis["irr"], "%"),
        ("ROI (Investitsiya qaytimi)", analysis["roi"], "%"),
        ("O'zini oqlash muddati", analysis["payback_period"], "yil"),
        ("Rentabellik indeksi (PI)", analysis["profitability_index"], ""),
        ("Yillik daromad", analysis["yillik_daromad"], "so'm"),
        ("Yillik xarajat", analysis["yillik_xarajat"], "so'm"),
        ("Yillik sof foyda", analysis["yillik_sof_foyda"], "so'm"),
    ]

    for label, val, unit in indicators:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.font = DATA_FONT
        if unit == "so'm":
            c.number_format = NUM_FORMAT
        ws.cell(row=row, column=3, value=unit).font = DATA_FONT
        row += 1

    row += 1

    # Xulosa
    xulosa = analysis["xulosa"]
    status = "✅ SAMARALI" if xulosa["loyiha_samarali"] else "⚠️ SAMARASIZ"
    ws.cell(row=row, column=1, value="LOYIHA BAHOSI:").font = Font(name='Times New Roman', bold=True, size=12)
    ws.cell(row=row, column=2, value=status).font = Font(
        name='Times New Roman', bold=True, size=12,
        color='27AE60' if xulosa["loyiha_samarali"] else 'E74C3C'
    )
    row += 2

    # Yillik jadval
    headers = ["Yil", "Daromad", "Xarajat", "Sof foyda", "Jami foyda"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        apply_cell_style(cell, HEADER_FONT, HEADER_FILL, Alignment(horizontal='center'), THIN_BORDER)
    row += 1

    for item in analysis["yillik_jadval"]:
        for i, val in enumerate([item["yil"], item["daromad"], item["xarajat"],
                                  item["sof_foyda"], item["jami_foyda"]], 1):
            cell = ws.cell(row=row, column=i, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if isinstance(val, float):
                cell.number_format = NUM_FORMAT
        row += 1

    return row + 2


def generate_dynamic_excel(template_path: str, output_path: str,
                           form_data: dict, kredit_natija: KreditNatija,
                           analysis: dict) -> str:
    """
    To'liq dinamik Excel yaratish:
    1. Asosiy ma'lumotlarni yozish
    2. Kredit jadvalini dinamik yaratish
    3. Moliyaviy tahlilni yozish
    """
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.worksheets[0]

    # === 1. Asosiy ma'lumotlar ===
    field_map = {
        "B16": "loyiha_nomi", "B18": "tashabbuskor", "B20": "manzil",
        "B22": "bank", "B24": "stir", "B26": "jshshir",
        "B28": "faoliyat", "B30": "mulk", "B32": "fio",
        "B34": "pasport", "B36": "berilgan_vaqti",
        "B38": "loyiha_qiymati", "B42": "oz_mablag",
        "B44": "kredit", "B46": "muddat", "B48": "imtiyoz", "B50": "foiz",
        "B56": "mahsulot", "B58": "hajm", "B60": "olchov", "B62": "narx",
        "B65": "direktor", "B67": "xodim", "B69": "yangi_xodim",
        "B72": "ishchi_oylik", "B74": "rahbar_oylik", "B76": "yangi_ishchi_oylik",
        "B81": "elektr", "B83": "gaz", "B85": "suv", "B87": "oqava",
    }

    for cell_ref, field_name in field_map.items():
        ws[cell_ref] = form_data.get(field_name, "")

    # Kredit summary
    ws["B90"] = kredit_natija.oylik_tolov
    ws["B91"] = kredit_natija.jami_tolov
    ws["B92"] = kredit_natija.jami_foiz

    # === 2. Kredit jadval sheet ===
    if "Kredit Jadvali" in wb.sheetnames:
        del wb["Kredit Jadvali"]
    ws_kredit = wb.create_sheet("Kredit Jadvali")
    next_row = write_kredit_jadval(ws_kredit, 1, kredit_natija)

    # === 3. Moliyaviy tahlil sheet ===
    if "Moliyaviy Tahlil" in wb.sheetnames:
        del wb["Moliyaviy Tahlil"]
    ws_analysis = wb.create_sheet("Moliyaviy Tahlil")
    write_financial_analysis(ws_analysis, 1, analysis)

    wb.save(output_path)
    return output_path

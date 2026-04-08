import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("e:/uzbiznesss_site/data.xlsx", data_only=False)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    for row in range(1, 10):
        row_vals = []
        for col in range(1, 15):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                row_vals.append(f"{get_column_letter(col)}{row}:{str(v)[:15]}")
        if row_vals:
            print("Row", row, ":", " | ".join(row_vals))


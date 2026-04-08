import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("e:/uzbiznesss_site/data.xlsx", data_only=False)

print("\n--- Sheet: Loans ---")
if "Loans" in wb:
    ws = wb["Loans"]
    for row in range(1, 20):
        row_vals = []
        for col in range(1, 8):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                row_vals.append(f"{get_column_letter(col)}{row}:{str(v)[:30]}")
        if row_vals:
            print("Row", row, ":", " | ".join(row_vals))
            
print("\n--- Sheet: Input (or whatever the main is) ---")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    v = ws['B16'].value if hasattr(ws['B16'], 'value') else None
    if v is not None:
        print(f"Sheet {sheet} B16: {v}")
